import GeneralFunctions as gf
import time, os, pyautogui, shutil, win32clipboard, openpyxl, pywinauto, json
from datetime import datetime, timedelta
from pywinauto.keyboard import send_keys
from openpyxl import Workbook

# Functions

def create_spec_directories(ui_self, main_path, main_directory_name, export_parent_directory, spec_directory_name, export_path, export_directory_name):
    if os.path.exists(main_path):
        gf.print_ui("Praeitos savaitės aplankas jau buvo sukurtas, informacija talpinam į ta patį aplanka.", ui_self)
        pass
    else:
        os.mkdir(main_path)
        gf.print_ui("Nurodytam take, sukurtas naujas aplankas pavadinimu: " + main_directory_name, ui_self)

    if os.path.exists(export_parent_directory):
        shutil.rmtree(export_parent_directory)
        
    os.mkdir(export_parent_directory)
    gf.print_ui("Aplanke [" + main_directory_name + "] sukurtas naujas aplankas: " + spec_directory_name, ui_self)

    os.mkdir(export_path)
    gf.print_ui("Aplanke [" + spec_directory_name + "] sukurtas naujas aplankas: " + export_directory_name, ui_self)

def find_same_images(image):
    if image[0][1] > image[1][1]:
        image_center = pyautogui.center(image[0])
        x, y = image_center
        pyautogui.click(x, y)
    elif image[0][1] < image[1][1]:
        image_center = pyautogui.center(image[1])
        x, y = image_center
        pyautogui.click(x, y)

def stock_export(group_name, filter_by_group, filter_by_sku, filter_by_title, main_process, export_path, scale): 
    gf.datagrid_search(filter_by_group, "Grupe", "Grupė", False, main_process)
    gf.datagrid_search(filter_by_sku, "Kodas", "Kodas", False, main_process)
    gf.datagrid_search(filter_by_title, "Pavadinimas", "Pavadinimas", False, main_process)
    send_keys("{ESC}")
    gf.locate_and_click(scale + "export-button.png", 0.8)
    gf.datagrid_search("", "Grupe", "Grupė", True, main_process)
    gf.datagrid_search("", "Kodas", "Kodas", True, main_process)
    gf.datagrid_search("", "Pavadinimas", "Pavadinimas", True, main_process)
    gf.clipboard_to_excel(group_name, "stock", export_path)

def pc_stock_export(ui_self, main_process, main_dlg, export_path, scale):
    gf.datagrid_search("k/komp", "Grupe", "Grupė", False, main_process)
    gf.datagrid_search("gamyba", "Operacija", "Operacija", False, main_process)
    send_keys("{ESC}")
    gf.locate_and_click(scale + "export-button.png", 0.8)
    
    win32clipboard.OpenClipboard()
    clipboard_data = win32clipboard.GetClipboardData()
    win32clipboard.EmptyClipboard()
    win32clipboard.CloseClipboard()
    clipboard_data = [i.split('\t') for i in clipboard_data.split('\r\n')]
    wb = Workbook()
    ws = wb.active
    for row, row_data in enumerate(clipboard_data, start=1):
        for col, cell_data in enumerate(row_data, start=1):
            ws.cell(row=row, column=col, value=cell_data)
    pc_count = ws.max_row - 2
    wb.close()
    gf.print_ui("Iš viso yra " + str(pc_count) + " surinktų kompiuterių.", ui_self)
    all_spec = Workbook()
    all_spec_sheet = all_spec.active
    next_row = 1

    gf.print_ui("Pradedama peržiurėti po viena...", ui_self)
    for row_i in range(pc_count):
        if row_i != 0:
            send_keys("{DOWN}")
        send_keys("{VK_APPS down}" "{VK_APPS up}")
        send_keys("{DOWN}")
        send_keys("{DOWN}")
        send_keys("{ENTER}")
        time.sleep(3) # Čia klaida buvo
        pc_dlg = main_dlg.child_window(class_name="TFormGamybosOperacija")
        pc_dlg.set_focus()
        pc_dlg.maximize()
        send_keys("%{ENTER}")
        
        for _ in range(60):
            a = pyautogui.locateOnScreen(scale + "konfiguracija.png", confidence=0.8)
            time.sleep(2)
            if a == None:
                continue
            else:
                try:
                    arrows = list(pyautogui.locateAllOnScreen(scale + "row-arrow.png", confidence=0.8))
                    time.sleep(3)
                    find_same_images(arrows)
                    break
                except Exception:
                    continue

        gf.datagrid_search("spec", "Pavadinimas", "Pavadinimas", False, main_process)

        export_buttons = list(pyautogui.locateAllOnScreen(scale + "export-button.png", confidence=0.8))
        time.sleep(3)
        find_same_images(export_buttons)
        
        time.sleep(1)
        win32clipboard.OpenClipboard()
        clipboard_data = win32clipboard.GetClipboardData()
        win32clipboard.EmptyClipboard()
        win32clipboard.CloseClipboard()
        clipboard_data = [i.split('\t') for i in clipboard_data.split('\r\n')]
        for row, row_data in enumerate(clipboard_data, start=next_row):
            for col, cell_data in enumerate(row_data, start=1):
                all_spec_sheet.cell(row=row, column=col, value=cell_data)
        if row_i != 0:
            all_spec_sheet.delete_rows(next_row)        
        next_row = all_spec_sheet.max_row
        pc_dlg.close()
        gf.print_ui("Peržiurėta " + str(row_i + 1) + " iš " + str(pc_count)+ " kompiuterių...", ui_self)
    
    max_col = all_spec_sheet.max_column
    for i in range(1, max_col + 1):
        cell_obj = all_spec_sheet.cell(row = 1, column = i)
        if cell_obj.value == "Kodas":
            column_of_sku = i
        elif cell_obj.value == "Pavadinimas":
            column_of_title = i
        elif cell_obj.value == "Kiekis":
            column_of_quantity = i
    
    rows_left = all_spec_sheet.max_row
    spec_rows = []
    for i in range(1, rows_left):
        sku_cell = all_spec_sheet.cell(row = i, column = column_of_sku).value
        title_cell = all_spec_sheet.cell(row = i, column = column_of_title).value
        qty_cell = all_spec_sheet.cell(row = i, column = column_of_quantity).value
        spec_rows.append([sku_cell, title_cell, qty_cell])

    all_spec_sheet.delete_rows(1,1000)

    total_spec_rows = []
    dup_free_set = set()
    for x in spec_rows:
        if tuple(x) not in dup_free_set:
            total_spec_rows.append(x)
            dup_free_set.add(tuple(x))

    for index, total_row in enumerate(total_spec_rows):
        if index == 0:
            continue
        total_quantity = 0
        for row in spec_rows:
            if total_row == row:
                total_quantity += 1
        total_spec_rows[index][2] = total_quantity
        
    for total_row in total_spec_rows:
        all_spec_sheet.append(tuple(total_row))

    all_spec.save(export_path + "\\" + "spec_in_pc" + ".xlsx")
    all_spec.close()

def insert_spec_column_values(export_sheet_obj, export_max_row, export_column, template_sheet_obj, template_column, template_alignment, value_to_float, select_row):
    for i in range(2, export_max_row + 1):
        cell_obj = export_sheet_obj.cell(row = i, column = export_column)
        template_cell_obj = template_sheet_obj.cell(row = i + select_row, column = template_column)
        template_cell_obj.alignment = openpyxl.styles.Alignment(horizontal=template_alignment)
        if value_to_float is True:
            if isinstance(cell_obj.value, str) and ',000' in cell_obj.value:
                template_cell_obj.value = float(cell_obj.value.replace(',000', ''))
            elif isinstance(cell_obj.value, str):
                template_cell_obj.value = float(cell_obj.value.replace(',', ''))
        else:
            template_cell_obj.value = cell_obj.value

def check_components(anchors, m_row, sheet_obj):
    filtered_rows = []
    for i in range(2, m_row + 1):
        cell_obj = sheet_obj.cell(row = i, column = 2)
        if isinstance(cell_obj.value, str):
            for anchor in anchors:
                if anchor.lower() in cell_obj.value.lower() and "kompiuter" not in cell_obj.value.lower():
                    sku_cell = sheet_obj.cell(row = i, column = 1).value
                    title_cell = sheet_obj.cell(row = i, column = 2).value
                    qty_cell = sheet_obj.cell(row = i, column = 3).value
                    filtered_rows.append([sku_cell, title_cell, qty_cell])
    return filtered_rows

def stock_data_to_report(component, spec_wb, export_path):
    path = export_path + "\\" + component + "_stock.xlsx"
    wb_obj = openpyxl.load_workbook(path)
    sheet_obj = wb_obj.active
    max_col = sheet_obj.max_column
    m_row = sheet_obj.max_row

    for i in range(1, max_col + 1):
        cell_obj = sheet_obj.cell(row = 1, column = i)
        if cell_obj.value == "Kodas":
            column_of_sku = i
        elif cell_obj.value == "Pavadinimas":
            column_of_title = i
        elif cell_obj.value == "Sand.likutis":
            column_of_quantity = i
    
    spec_sheet = spec_wb[component.upper()]

    insert_spec_column_values(sheet_obj, m_row, column_of_sku, spec_sheet, 1, "left", False, 0)
    insert_spec_column_values(sheet_obj, m_row, column_of_title, spec_sheet, 2, "left", False, 0)
    insert_spec_column_values(sheet_obj, m_row, column_of_quantity, spec_sheet, 3, "center", True, 0)
    wb_obj.close()

def pc_stock_data_to_report(component, spec_wb, export_path, vga_filter_list, mb_filter_list, psu_filter_list, case_filter_list, cooler_filter_list, ssd_filter_list):
    path = export_path + "\\" + "spec_in_pc.xlsx"
    wb_obj = openpyxl.load_workbook(path)
    sheet_obj = wb_obj.active
    m_row = sheet_obj.max_row

    spec_sheet = spec_wb[component.upper()]
    spec_m_row = spec_sheet.max_row

    if component == "vga":
        filtered_rows = check_components(vga_filter_list, m_row, sheet_obj)
    elif component == "mb":
        filtered_rows = check_components(mb_filter_list, m_row, sheet_obj)
    elif component == "psu":
        filtered_rows = check_components(psu_filter_list, m_row, sheet_obj)
    elif component == "case":
        filtered_rows = check_components(case_filter_list, m_row, sheet_obj)
    elif component == "cooler":
        filtered_rows = check_components(cooler_filter_list, m_row, sheet_obj)
    elif component == "ssd":
        filtered_rows = check_components(ssd_filter_list, m_row, sheet_obj)

    spec_sheet.delete_rows(spec_m_row)
    for row in filtered_rows:
        spec_sheet.append(tuple(row))
    
    rows_left = spec_sheet.max_row
    spec_rows = []
    for i in range(2, rows_left + 1):
        sku_cell = spec_sheet.cell(row = i, column = 1).value
        title_cell = spec_sheet.cell(row = i, column = 2).value
        qty_cell = float(spec_sheet.cell(row = i, column = 3).value)
        spec_rows.append([sku_cell, title_cell, qty_cell])

    spec_sheet.delete_rows(2,1000)

    final_list = []
    for idx_row_1, row_1 in enumerate(spec_rows):
        sku = row_1[0]
        title = row_1[1]
        qty = row_1[2]
        for idx_row_2, row_2 in enumerate(spec_rows):
            if idx_row_1 != idx_row_2:
                if row_1[0] in row_2[0] and row_1[1] in row_2[1]:
                    qty = row_1[2] + row_2[2]
                    spec_rows.pop(idx_row_2)
        final_list.append([sku, title, qty])
        
    for idx, row in enumerate(final_list, start=2):
        spec_sheet.append(tuple(row))
        qty_column = spec_sheet.cell(row = idx, column = 3)
        qty_column.alignment = openpyxl.styles.Alignment(horizontal="center")

def main(ui_self, user_name, user_password):
    # Variables - Date and Time variables
    today_date = datetime.now()
    prior_week_end = (datetime.now() - timedelta(days=((datetime.now().isoweekday()) % 7))).strftime("%Y.%m.%d")
    prior_week_start = ((datetime.now() - timedelta(days=((datetime.now().isoweekday()) % 7))) - timedelta(days=6)).strftime("%Y.%m.%d")
    year, week_num, day_of_week = today_date.isocalendar()

    # Variables - Variables for files and directories
    default_path = "C:/Dokumentai/Automatiškai sugeneruoti duomenys"
    f = open("data/DirectoryData.json", encoding='utf-8')
    f_data = json.load(f)
    directories = f_data["directories"]
    main_parent_directory = directories["report_directory"]
    f.close()
    main_directory_name = str(prior_week_start) + " - " + str(prior_week_end)
    main_path = os.path.join(main_parent_directory, main_directory_name)
    spec_directory_name = "SPEC likučiai"
    export_directory_name = "Eksportas SPEC likučių"
    export_parent_directory = os.path.join(main_path, spec_directory_name)
    export_path = os.path.join(export_parent_directory, export_directory_name)

    # Component data variables
    vga_filter_list = []
    mb_filter_list = []
    psu_filter_list = []
    case_filter_list = []
    cooler_filter_list = []
    ssd_filter_list = []

    f = open("data/SPECComponentSettings.json",)
    f_data = json.load(f)
    for filter in f_data["VGA"]:
        vga_filter_list.append(filter)
    for filter in f_data["MB"]:
        mb_filter_list.append(filter)
    for filter in f_data["PSU"]:
        psu_filter_list.append(filter)
    for filter in f_data["CASE"]:
        case_filter_list.append(filter)
    for filter in f_data["COOLER"]:
        cooler_filter_list.append(filter)
    for filter in f_data["SSD"]:
        ssd_filter_list.append(filter)
    f.close()

    # Step 1: Connects to the Application, creates directories and calibrates the main window to user monitor settings
    if week_num < 10:
        week_num = "0" + str(week_num)
    else:
        week_num = str(week_num)

    start_time = time.time()
    main_process, main_dlg = gf.connect(ui_self, user_name, user_password)

    if default_path == main_parent_directory and not os.path.isdir(default_path):
        os.makedirs(default_path)
    create_spec_directories(ui_self, main_path, main_directory_name, export_parent_directory, spec_directory_name, export_path, export_directory_name)
    pywinauto.mouse.move(coords=(10,10))
    time.sleep(1)
    gf.close_window(ui_self, "TFormGamybosOperacija", main_dlg)
    gf.close_window(ui_self, "TFormEinamiejiVertybiuLikuciai", main_dlg)
    gf.close_window(ui_self, "TFormKorteliuLikuciai", main_dlg)

    scale = str(gf.find_current_dpi(ui_self, "meniu.png", 0.8))

    # Step 2: Exports SPEC stock from the program
    gf.print_ui("Pradedama eksportuoti SPEC likučius iš įmonės programos...", ui_self)
    gf.navigate_menu(ui_self, scale + "ataskaitos.png", 0.8, scale + "vertybiu-likuciai.png", 0.8, scale + "faktiniai-likuciai.png", 0.8, main_dlg)
    stock_dlg = main_dlg.child_window(class_name="TFormEinamiejiVertybiuLikuciai")
    stock_dlg.set_focus()
    stock_dlg.minimize()
    stock_dlg.maximize()
    gf.refresh_data(scale)

    stock_export("vga", "kk/vga", "", "spec", main_process, export_path, scale)
    gf.print_ui("SPEC vaizdo plokščių (VGA) likučiai eksportuoti...", ui_self)

    stock_export("mb", "kk/mb", "", "spec", main_process, export_path, scale)
    gf.print_ui("SPEC pagrindinių plokščių (MB) likučiai eksportuoti...", ui_self)

    stock_export("psu", "kk/psu", "", "spec", main_process, export_path, scale)
    gf.print_ui("SPEC maitinimo blokų (PSU) likučiai eksportuoti...", ui_self)

    stock_export("case", "kk/case", "", "spec", main_process, export_path, scale)
    gf.print_ui("SPEC korpusų (CASE) likučiai eksportuoti...", ui_self)

    stock_export("cooler", "kk/cool", "", "spec", main_process, export_path, scale)
    gf.print_ui("SPEC procesoriaus aušintuvų (COOLER) likučiai eksportuoti...", ui_self)

    stock_export("ssd", "kk/hdd", "", "spec", main_process, export_path, scale)
    gf.print_ui("SPEC kietųjų diskų (SSD) likučiai eksportuoti...", ui_self)

    stock_dlg.close()
    gf.print_ui("Pabaigta eksportuoti SPEC likučius.", ui_self)

    # Step 3: Exports SPEC stock in PC from the program
    gf.print_ui("Toliau pradedam likučius ištraukinėt iš kompiuterių...", ui_self)
    gf.navigate_menu(ui_self, scale + "ataskaitos.png", 0.8, scale + "vertybiu-likuciai.png", 0.8, scale + "partiju-likuciai.png", 0.8, main_dlg)
    pc_stock_dlg = main_dlg.child_window(class_name="TFormKorteliuLikuciai")
    pc_stock_dlg.set_focus()
    pc_stock_dlg.minimize()
    pc_stock_dlg.maximize()
    gf.refresh_data(scale)
    pc_stock_export(ui_self, main_process, main_dlg, export_path, scale)
    pc_stock_dlg.close()
    gf.print_ui("Pabaigta eksportuoti likučius.", ui_self)
    gf.print_ui("Visi duomenys iš įmonės programos eksportuoti sėkmingai.", ui_self)

    # Step 4: Creates SPEC STOCK report
    gf.print_ui("Pradedama ruošti SPEC ataskaita...", ui_self)
    spec_wb = openpyxl.load_workbook("templates/spec_stock_example.xlsx")
    stock_data_to_report("vga", spec_wb, export_path)
    stock_data_to_report("mb", spec_wb, export_path)
    stock_data_to_report("psu", spec_wb, export_path)
    stock_data_to_report("case", spec_wb, export_path)
    stock_data_to_report("cooler", spec_wb, export_path)
    stock_data_to_report("ssd", spec_wb, export_path)
    pc_stock_data_to_report("vga", spec_wb, export_path, vga_filter_list, mb_filter_list, psu_filter_list, case_filter_list, cooler_filter_list, ssd_filter_list)
    pc_stock_data_to_report("mb", spec_wb, export_path, vga_filter_list, mb_filter_list, psu_filter_list, case_filter_list, cooler_filter_list, ssd_filter_list)
    pc_stock_data_to_report("psu", spec_wb, export_path, vga_filter_list, mb_filter_list, psu_filter_list, case_filter_list, cooler_filter_list, ssd_filter_list)
    pc_stock_data_to_report("case", spec_wb, export_path, vga_filter_list, mb_filter_list, psu_filter_list, case_filter_list, cooler_filter_list, ssd_filter_list)
    pc_stock_data_to_report("cooler", spec_wb, export_path, vga_filter_list, mb_filter_list, psu_filter_list, case_filter_list, cooler_filter_list, ssd_filter_list)
    pc_stock_data_to_report("ssd", spec_wb, export_path, vga_filter_list, mb_filter_list, psu_filter_list, case_filter_list, cooler_filter_list, ssd_filter_list)

    spec_wb.save(export_parent_directory + "\\" + "SPEC_stock_" + str(year) + "_W" + week_num + ".xlsx")
    spec_wb.close()
    
    gf.print_ui("Pabaigta ruošti SPEC ataskaita.", ui_self)
    end_time = time.time()
    seconds_spent_executing = end_time - start_time
    executing_time = timedelta(seconds=seconds_spent_executing)
    gf.print_ui("Visi SPEC savaitinės ataskaitos modulio darbai pabaigti per: [" + str(executing_time) + "]", ui_self)

if __name__ == "__main__":
    main(None, "", "")