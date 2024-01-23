import GeneralFunctions as gf
import time, os, pyautogui, shutil, win32clipboard, openpyxl, pywinauto, json
from datetime import datetime, timedelta
from pywinauto.keyboard import send_keys
from openpyxl import Workbook
from openpyxl.styles import Font, Color, Alignment, PatternFill

# Functions

def create_spec_monthly_directories(ui_self, main_path, main_directory_name, export_parent_directory, monthly_directory, export_path, export_directory_name):
    if os.path.exists(main_path):
        gf.print_ui("SPEC mėnesinių ataskaitų aplankas jau yra sukurtas, informacija talpinam į ta patį aplanka.", ui_self)
        pass
    else:
        os.mkdir(main_path)
        gf.print_ui("Nurodytam take, sukurtas naujas aplankas pavadinimu: " + main_directory_name, ui_self)

    if os.path.exists(export_parent_directory):
        shutil.rmtree(export_parent_directory)
        
    os.mkdir(export_parent_directory)
    gf.print_ui("Aplanke [" + main_directory_name + "] sukurtas naujas aplankas: " + monthly_directory, ui_self)

    os.mkdir(export_path)
    gf.print_ui("Aplanke [" + monthly_directory + "] sukurtas naujas aplankas: " + export_directory_name, ui_self)

def clipboard_to_excel_without_save(wb):
    win32clipboard.OpenClipboard()
    clipboard_data = win32clipboard.GetClipboardData()
    win32clipboard.CloseClipboard()
    clipboard_data = clipboard_data.replace('\t ', '\t')
    clipboard_data = [i.split('\t') for i in clipboard_data.split('\r\n')]
    ws = wb.active
    first_row = ws.max_row
    for row, row_data in enumerate(clipboard_data, start=ws.max_row):
        for col, cell_data in enumerate(row_data, start=1):
            ws.cell(row=row, column=col, value=cell_data)
    
    if (first_row != 1):
        ws.delete_rows(first_row)

def stock_export(group_name, filter_by_group, filter_by_sku, filter_by_title, main_process, export_path, scale): 
    gf.datagrid_search(filter_by_group, "Grupe", "Grupė", False, main_process)
    gf.datagrid_search(filter_by_sku, "Kodas", "Kodas", False, main_process)
    gf.datagrid_search(filter_by_title, "Pavadinimas", "Pavadinimas", False, main_process)
    send_keys("{ESC}")
    gf.locate_and_click(scale + "export-button.png", 0.8)
    gf.datagrid_search("", "Grupe", "Grupė", True, main_process)
    gf.datagrid_search("", "Kodas", "Kodas", True, main_process)
    gf.datagrid_search("", "Pavadinimas", "Pavadinimas", True, main_process)
    gf.clipboard_to_excel(group_name, "stock", export_path)

def sales_export(company, filter_by_group, filter_by_sku, filter_by_title, main_process, export_path, scale):
    gf.datagrid_search(filter_by_group, "Grupe", "Grupė", False, main_process)
    gf.datagrid_search(filter_by_sku, "Pr.kodas", "Pr.kodas", False, main_process)
    gf.datagrid_search(filter_by_title, "Pr.pavad.", "Pr.pavad.", False, main_process)
    send_keys("{ESC}")
    gf.locate_and_click(scale + "export-button.png", 0.8)
    gf.datagrid_search("", "Grupe", "Grupė", True, main_process)
    gf.datagrid_search("", "Pr.kodas", "Pr.kodas", True, main_process)
    gf.datagrid_search("", "Pr.pavad.", "Pr.pavad.", True, main_process)
    
    gf.clipboard_to_excel(company.lower(), "sale_out", export_path)

def pc_sales_export(company, filter_by_sku, filter_by_title, main_process, export_path, scale):
    gf.datagrid_search(filter_by_sku, "ˇaliava", "Žaliava", False, main_process)
    gf.datagrid_search(filter_by_title, "ˇ.pavadinimas", "Ž.pavadinimas", False, main_process)
    send_keys("{ESC}")
    gf.locate_and_click(scale + "export-button.png", 0.8)
    gf.datagrid_search("", "ˇaliava", "Žaliava", True, main_process)
    gf.datagrid_search("", "ˇ.pavadinimas", "Ž.pavadinimas", True, main_process)

def check_components(anchors, m_row, sheet_obj):
    filtered_rows = []
    for i in range(2, m_row + 1):
        cell_obj = sheet_obj.cell(row = i, column = 2)
        if isinstance(cell_obj.value, str):
            for anchor in anchors:
                if anchor.lower() in cell_obj.value.lower() and "kompiuter" not in cell_obj.value.lower():
                    sku_cell = sheet_obj.cell(row = i, column = 1).value
                    title_cell = sheet_obj.cell(row = i, column = 2).value
                    qty_cell = sheet_obj.cell(row = i, column = 3).value
                    filtered_rows.append([sku_cell, title_cell, qty_cell])
    return filtered_rows

def export_without_group(company, filter_by_sku, filter_by_title, main_process, export_path, scale):
    gf.datagrid_search(filter_by_sku, "Prekes kodas", "Prekės kodas", False, main_process)
    gf.datagrid_search(filter_by_title, "Prekes pavadinimas", "Prekės pavadinimas", False, main_process)
    send_keys("{ESC}")
    gf.locate_and_click(scale + "export-button.png", 0.8)
    gf.datagrid_search("", "Prekes kodas", "Prekės kodas", True, main_process)
    gf.datagrid_search("", "Prekes pavadinimas", "Prekės pavadinimas", True, main_process)

def check_if_by_group(ui_self, scale):
    for _ in range(15):
        send_keys("{UP}")

    check_dropdown = pyautogui.locateOnScreen(scale + "pagal-prekiu-grupes.png", confidence=0.8)
    if check_dropdown == None:
        gf.print_ui("Nerado pagal prekių grupes pasirinkimo...", ui_self)
        for _ in range(15):
            check_dropdown = pyautogui.locateOnScreen(scale + "pagal-prekiu-grupes.png", confidence=0.8)
            if check_dropdown == None:
                send_keys("{DOWN}")
            else:
                gf.locate_and_click(scale + "pagal-prekiu-grupes.png", 0.8)
                break
    else:
        gf.locate_and_click(scale + "pagal-prekiu-grupes.png", 0.8)

def sale_in_append(export_path, file):
    path = export_path + "\\" + file
    wb_obj = openpyxl.load_workbook(path)
    sheet_obj = wb_obj.active
    
    max_col = sheet_obj.max_column
    for i in range(1, max_col + 1):
        cell_obj = sheet_obj.cell(row = 1, column = i)
        if cell_obj.value == "Prekes kodas" or cell_obj.value == "Prekės kodas":
            column_of_sku = i
        elif cell_obj.value == "Prekes pavadinimas" or cell_obj.value == "Prekės pavadinimas":
            column_of_title = i
        elif cell_obj.value == "Kiekis":
            column_of_quantity = i

    rows = sheet_obj.max_row
    sale_in_data = []
    for i in range(1, rows):
        sku_cell = sheet_obj.cell(row = i, column = column_of_sku).value
        title_cell = sheet_obj.cell(row = i, column = column_of_title).value
        if i == 1:
            qty_cell = sheet_obj.cell(row = i, column = column_of_quantity).value
        else:
            qty_value = sheet_obj.cell(row = i, column = column_of_quantity).value
            if ',000' in qty_value:
                qty_value = float(qty_value.replace(',000', ''))
                
            qty_cell = int(float(qty_value))
        sale_in_data.append([sku_cell, title_cell, qty_cell])

    sheet_obj.delete_rows(1,1000)
    final_list = []

    for idx_row_1, row_1 in enumerate(sale_in_data):
        if row_1[0] != "":
            for idx_row_2, row_2 in enumerate(sale_in_data):
                if idx_row_1 != idx_row_2:
                    if row_1[0] == row_2[0] and row_1[1] == row_2[1]:
                        row_1[2] = row_1[2] + row_2[2]
                        row_2[0] = ""
                        row_2[1] = ""
                        row_2[2] = 0
            final_list.append([row_1[0], row_1[1], row_1[2]])
        else:
            continue
    
    for row in final_list:
        sheet_obj.append(row)
    
    wb_obj.save(export_path + "\\" + file)

def export_to_sheet(wb_obj, report_group, sale_date, stock_date, export_path, manufacturers, vga_filter_list, mb_filter_list, psu_filter_list, cooler_filter_list, case_filter_list, ssd_filter_list):
    theme_color = Color(rgb='00DDEBF7')
    theme_fill = PatternFill(patternType='solid', fgColor=theme_color)
    wb_obj.create_sheet(report_group)
    sheet_obj = wb_obj[report_group]
    sheet_obj.column_dimensions['A'].width = 40
    sheet_obj.column_dimensions['B'].width = 50
    sheet_obj.column_dimensions['C'].width = 25
    sheet_obj['A1'].fill = theme_fill
    sheet_obj['B1'].fill = theme_fill
    sheet_obj['C1'].fill = theme_fill
    sheet_obj['A1'].font = Font(size=8, bold=True)
    sheet_obj['B1'].font = Font(size=8, bold=True)
    sheet_obj['C1'].font = Font(size=8, bold=True)
    sheet_obj['A1'].alignment = Alignment(horizontal="center", vertical="center")
    sheet_obj['B1'].alignment = Alignment(horizontal="center", vertical="center")
    sheet_obj['C1'].alignment = Alignment(horizontal="center", vertical="center")
    sheet_obj['B1'].value = "Retail Name"
    
    
    ex_file_path = export_path + "\\" + report_group + ".xlsx"
    ex_wb = openpyxl.load_workbook(ex_file_path)
    ex_sheet = ex_wb.active
    ex_max_col = ex_sheet.max_column
    
    is_stock = False
    if "sale_in" in report_group:
        sheet_obj['C1'].value = "Sale in qty " + sale_date
        ex_m_row = ex_sheet.max_row
    elif "with_PC_sale_out" in report_group:
        sheet_obj['C1'].value = "Sale out qty " + sale_date
        ex_m_row = ex_sheet.max_row - 1
    elif "sale_out" in report_group:
        sheet_obj['C1'].value = "Sale out qty " + sale_date
        ex_m_row = ex_sheet.max_row - 2
    elif "stock" in report_group:
        sheet_obj['C1'].value = "Stock qty " + stock_date
        ex_m_row = ex_sheet.max_row - 1
        is_stock = True
    
    is_spec_other = False
    if "SPEC" in report_group:
        is_spec_other = True


    for i in range(1, ex_max_col + 1):
        ex_cell_obj = ex_sheet.cell(row = 1, column = i)
        if ex_cell_obj.value == "Prekės kodas" or ex_cell_obj.value == "Prekes kodas" or ex_cell_obj.value == "Pr.kodas" or ex_cell_obj.value == "Kodas" or ex_cell_obj.value == "ˇaliava" or ex_cell_obj.value == "Žaliava":
            column_of_sku = i
        elif ex_cell_obj.value == "Prekės pavadinimas" or ex_cell_obj.value == "Prekes pavadinimas" or ex_cell_obj.value == "Pr.pavad." or ex_cell_obj.value == "Pavadinimas" or ex_cell_obj.value == "ˇ.pavadinimas" or ex_cell_obj.value == "Ž.pavadinimas":
            column_of_title = i
        elif ex_cell_obj.value == "Kiekis" or ex_cell_obj.value == "Sand.likutis" or ex_cell_obj.value == "ˇ.kiekis" or ex_cell_obj.value == "Ž.kiekis":
            column_of_quantity = i
            
    list_by_manufacturer = []
    active_manufacturers = []

    for i in range(2, ex_m_row + 1):
        cell_obj = ex_sheet.cell(row = i, column = column_of_sku)
        sku = cell_obj.value
        cell_obj = ex_sheet.cell(row = i, column = column_of_title)
        title = cell_obj.value
        cell_obj = ex_sheet.cell(row = i, column = column_of_quantity)
        quantity = cell_obj.value

        is_categorized = False
        for manufacturer in manufacturers:
            if isinstance(title, str):
                if manufacturer.lower() in title.lower():
                    is_categorized = True
                    list_by_manufacturer.append([manufacturer, sku, title, quantity])
                    active_manufacturers.append(manufacturer)
        
        if is_categorized is False:
            list_by_manufacturer.append(["Nepriskirtas gamintojas", sku, title, quantity])
            active_manufacturers.append("Nepriskirtas gamintojas")
    
    
    active_manufacturers = set(active_manufacturers)
    if is_stock:
        active_manufacturers = {"SPEC", "Nepriskirtas gamintojas"}

    if is_spec_other:
        group_by_type("MONITOR", sheet_obj, theme_fill, list_by_manufacturer, vga_filter_list, mb_filter_list, psu_filter_list, cooler_filter_list, case_filter_list, ssd_filter_list, True)
        group_by_type("CASE", sheet_obj, theme_fill, list_by_manufacturer, vga_filter_list, mb_filter_list, psu_filter_list, cooler_filter_list, case_filter_list, ssd_filter_list, False)
        group_by_type("PSU", sheet_obj, theme_fill, list_by_manufacturer, vga_filter_list, mb_filter_list, psu_filter_list, cooler_filter_list, case_filter_list, ssd_filter_list, False)
        group_by_type("COOLING", sheet_obj, theme_fill, list_by_manufacturer, vga_filter_list, mb_filter_list, psu_filter_list, cooler_filter_list, case_filter_list, ssd_filter_list, False)
        group_by_type("SSD", sheet_obj, theme_fill, list_by_manufacturer, vga_filter_list, mb_filter_list, psu_filter_list, cooler_filter_list, case_filter_list, ssd_filter_list, False)
        group_by_type("GAMING GEAR", sheet_obj, theme_fill, list_by_manufacturer, vga_filter_list, mb_filter_list, psu_filter_list, cooler_filter_list, case_filter_list, ssd_filter_list, False)
    else:
        current_row = sheet_obj.max_row + 1
        for active_manufacturer in active_manufacturers:
            manufacturer_cell = sheet_obj.cell(row = current_row, column = 1)
            empty_cell_1 = sheet_obj.cell(row = current_row, column = 2)
            empty_cell_2 = sheet_obj.cell(row = current_row, column = 3)
            manufacturer_cell.alignment = Alignment(horizontal="left", vertical="center")
            manufacturer_cell.value = active_manufacturer
            
            if active_manufacturer == "Nepriskirtas gamintojas":
                theme_color = Color(rgb='00F2DCDB')
                theme_fill = PatternFill(patternType='solid', fgColor=theme_color)
            else:
                theme_color = Color(rgb='00DDEBF7')
                theme_fill = PatternFill(patternType='solid', fgColor=theme_color)

            manufacturer_cell.fill = theme_fill
            manufacturer_cell.font = Font(size=15, bold=True)
            empty_cell_1.fill = theme_fill
            empty_cell_2.fill = theme_fill

            
            for row in list_by_manufacturer:
                if row[0] == active_manufacturer and row[2] != "":
                    current_row += 1
                    sku_cell_obj = sheet_obj.cell(row = current_row, column = 1)
                    sku_cell_obj.alignment = Alignment(horizontal="left", vertical="center")
                    sku_cell_obj.font = Font(size=12)
                    sku_cell_obj.value = row[1]

                    title_cell_obj = sheet_obj.cell(row = current_row, column = 2)
                    title_cell_obj.alignment = Alignment(horizontal="left", vertical="center")
                    title_cell_obj.font = Font(size=12)
                    title_cell_obj.value = row[2]

                    qty_cell_obj = sheet_obj.cell(row = current_row, column = 3)
                    qty_cell_obj.alignment = Alignment(horizontal="center", vertical="center")
                    qty_cell_obj.font = Font(size=12)
                    try:
                        qty_cell_obj.value = int(float(row[3]))
                    except Exception:
                        qty_cell_obj.value = row[3] 
            
            current_row += 2
    
    ex_wb.close()

def group_by_type(group, sheet_obj, theme_fill, list_by_manufacturer, vga_filter_list, mb_filter_list, psu_filter_list, cooler_filter_list, case_filter_list, ssd_filter_list, first_group):
    if first_group:
        current_row = sheet_obj.max_row + 1
    else:
        current_row = sheet_obj.max_row + 2

    active_manufacturer = "SPEC"
    manufacturer_cell = sheet_obj.cell(row = current_row, column = 1)
    empty_cell_1 = sheet_obj.cell(row = current_row, column = 2)
    empty_cell_2 = sheet_obj.cell(row = current_row, column = 3)
    manufacturer_cell.alignment = Alignment(horizontal="left", vertical="center")
    manufacturer_cell.value = group
    manufacturer_cell.fill = theme_fill
    manufacturer_cell.font = Font(size=15, bold=True)
    empty_cell_1.fill = theme_fill
    empty_cell_2.fill = theme_fill
    for row in list_by_manufacturer:
        if row[0] == active_manufacturer:
            is_vga_or_mb = False
            is_psu = False
            is_cooler = False
            is_case = False
            is_monitor = False
            is_ssd = False
            is_gg = False

            for vga in vga_filter_list:
                if vga.lower() in row[2].lower():
                    is_vga_or_mb = True
            
            for mb in mb_filter_list:
                if mb.lower() in row[2].lower():
                    is_vga_or_mb = True
            
            for psu in psu_filter_list:
                if psu.lower() in row[2].lower():
                    is_psu = True
            
            for cooler in cooler_filter_list:
                if cooler.lower() in row[2].lower():
                    is_cooler = True

            for case in case_filter_list:
                if case.lower() in row[2].lower():
                    is_case = True

            for ssd in ssd_filter_list:
                if ssd.lower() in row[2].lower():
                    is_ssd = True

            if "monitor" in row[2].lower() or "optix" in row[2].lower() or "lcd" in row[2].lower():
                is_monitor = True

            if (not is_vga_or_mb) and (not is_psu) and (not is_cooler) and (not is_case) and (not is_ssd) and (not is_monitor):
                is_gg = True

            if is_vga_or_mb:
                continue
            elif (group == "MONITOR" and is_monitor) or (group == "CASE" and is_case) or (group == "PSU" and is_psu) or (group == "COOLING" and is_cooler) or (group == "SSD" and is_ssd) or (group == "GAMING GEAR" and is_gg):
                current_row += 1
                sku_cell_obj = sheet_obj.cell(row = current_row, column = 1)
                sku_cell_obj.alignment = Alignment(horizontal="left", vertical="center")
                sku_cell_obj.font = Font(size=12)
                sku_cell_obj.value = row[1]

                title_cell_obj = sheet_obj.cell(row = current_row, column = 2)
                title_cell_obj.alignment = Alignment(horizontal="left", vertical="center")
                title_cell_obj.font = Font(size=12)
                title_cell_obj.value = row[2]

                qty_cell_obj = sheet_obj.cell(row = current_row, column = 3)
                qty_cell_obj.alignment = Alignment(horizontal="center", vertical="center")
                qty_cell_obj.font = Font(size=12)
                try:
                    qty_cell_obj.value = int(float(row[3]))
                except Exception:
                    qty_cell_obj.value = row[3]

def main(ui_self, user_name, user_password):
    # Variables - Date and Time variables
    today_date = datetime.now().strftime("%Y.%m.%d")
    prior_month_end = datetime.now().replace(day=1) - timedelta(days=1)
    prior_month_start = prior_month_end.replace(day=1)

    if prior_month_start.month == 1:
        monthly_directory = str(today_date) + " už Sausio mėnesį"
    elif prior_month_start.month == 2:
        monthly_directory = str(today_date) + " už Vasario mėnesį"
    elif prior_month_start.month == 3:
        monthly_directory = str(today_date) + " už Kovo mėnesį"
    elif prior_month_start.month == 4:
        monthly_directory = str(today_date) + " už Balandžio mėnesį"
    elif prior_month_start.month == 5:
        monthly_directory = str(today_date) + " už Gegužės mėnesį"
    elif prior_month_start.month == 6:
        monthly_directory = str(today_date) + " už Biržėlio mėnesį"
    elif prior_month_start.month == 7:
        monthly_directory = str(today_date) + " už Liepos mėnesį"
    elif prior_month_start.month == 8:
        monthly_directory = str(today_date) + " už Rugpjūčio mėnesį"
    elif prior_month_start.month == 9:
        monthly_directory = str(today_date) + " už Rugsėjo mėnesį"
    elif prior_month_start.month == 10:
        monthly_directory = str(today_date) + " už Spalio mėnesį"
    elif prior_month_start.month == 11:
        monthly_directory = str(today_date) + " už Lapkričio mėnesį"
    elif prior_month_start.month == 12:
        monthly_directory = str(today_date) + " už Gruoždio mėnesį"

    
    translate_date_begin_lt = prior_month_start.strftime("%Y-%m-%d")
    translate_date_end_lt = prior_month_end.strftime("%Y-%m-%d")
    
    prior_month_start = prior_month_start.strftime("%Y.%m.%d")
    prior_month_end = prior_month_end.strftime("%Y.%m.%d")
    
    translate_date_begin = prior_month_start[5] + prior_month_start[6] + prior_month_start[8] + prior_month_start[9] + prior_month_start[0] + prior_month_start[1] + prior_month_start[2] + prior_month_start[3]
    translate_date_end = prior_month_end[5] + prior_month_end[6] + prior_month_end[8] + prior_month_end[9] + prior_month_end[0] + prior_month_end[1] + prior_month_end[2] + prior_month_end[3]


    # Variables - Variables for files and directories
    
    default_path = "C:/Dokumentai/Automatiškai sugeneruoti duomenys"
    f = open("data/DirectoryData.json", encoding='utf-8')
    f_data = json.load(f)
    directories = f_data["directories"]
    main_parent_directory = directories["report_directory"]
    f.close()
    main_directory_name = "SPEC mėnesinės ataskaitos"
    main_path = os.path.join(main_parent_directory, main_directory_name)
    export_directory_name = "SPEC Eksportas"
    export_parent_directory = os.path.join(main_path, monthly_directory)
    export_path = os.path.join(export_parent_directory, export_directory_name)

    # Manufacturer data
    manufacturers = []
    f = open("data/ManufacturerData.json",)
    f_data = json.load(f)
    for manufacturer in f_data["manufacturers"]:
        manufacturers.append(manufacturer)
    f.close()

    # Component data variables
    vga_filter_list = []
    mb_filter_list = []
    psu_filter_list = []
    case_filter_list = []
    cooler_filter_list = []
    ssd_filter_list = []

    f = open("data/SPECComponentSettings.json",)
    f_data = json.load(f)
    for filter in f_data["VGA"]:
        vga_filter_list.append(filter)
    for filter in f_data["MB"]:
        mb_filter_list.append(filter)
    for filter in f_data["PSU"]:
        psu_filter_list.append(filter)
    for filter in f_data["CASE"]:
        case_filter_list.append(filter)
    for filter in f_data["COOLER"]:
        cooler_filter_list.append(filter)
    for filter in f_data["SSD"]:
        ssd_filter_list.append(filter)
    f.close()

    # Step 1: Connects to the Application, creates directories and calibrates the main window to user monitor settings

    start_time = time.time()
    main_process, main_dlg = gf.connect(ui_self, user_name, user_password)

    if default_path == main_parent_directory and not os.path.isdir(default_path):
        os.makedirs(default_path)
    create_spec_monthly_directories(ui_self, main_path, main_directory_name, export_parent_directory, monthly_directory, export_path, export_directory_name)
    pywinauto.mouse.move(coords=(10,10))
    time.sleep(1)
    gf.close_window(ui_self, "TFormEinamiejiVertybiuLikuciai", main_dlg)
    gf.close_window(ui_self, "TForm_PirkimuKnyga", main_dlg)
    gf.close_window(ui_self, "TFormPrekiuPardavimaiPagalPartnerius", main_dlg)
    gf.close_window(ui_self, "TFormZaliavuPanaudojimasGaminiuose", main_dlg)

    scale = str(gf.find_current_dpi(ui_self, "meniu.png", 0.8))

    gf.print_ui("Tikrinama įmonės programos versija...", ui_self)
    time.sleep(2)
    try:
        window_text = []
        for child in main_dlg.children():
            window_text += child.texts()
        version = None
        if "\tv. 4.0.48.248" in window_text:
            gf.print_ui("Aptikta v. 4.0.48.248 įmonės programos versija. Versija palaikoma.", ui_self)
            version = "v. 4.0.48.248"
        elif "\tv. 4.0.48.243" in window_text:
            gf.print_ui("Aptikta v. 4.0.48.243 įmonės programos versija. Versija palaikoma.", ui_self)
            version = "v. 4.0.48.243"
        else:
            gf.print_ui("Įmonės programos versija nėra palaikoma, darbai bus atliekami kaip v. 4.0.48.248 versijai.", ui_self)
    except Exception as e: gf.print_ui(e, ui_self)

    # Step 2: Exports sale out

    gf.print_ui("Pradedama eksportuoti pardavimus iš įmonės programos...", ui_self)

    gf.navigate_menu(ui_self, scale + "ataskaitos.png", 0.8, scale + "prekybos-analize.png", 0.9, scale + "prekiu-pardavimai.png", 0.9, main_dlg)
    sales_dlg = main_dlg.child_window(class_name="TFormPrekiuPardavimaiPagalPartnerius")
    sales_dlg.set_focus()
    time.sleep(2)
    sales_dlg.minimize()
    sales_dlg.maximize()

    if version == "v. 4.0.48.243":
        textbox_start = sales_dlg.Edit11
        textbox_end = sales_dlg.Edit10
    else:
        textbox_start = sales_dlg.Edit13
        textbox_end = sales_dlg.Edit12

    combobox_order = sales_dlg.child_window(title="Pagal partnerius", class_name="TRxDBLookupCombo")
    combobox_type = sales_dlg.child_window(best_match="Sumin")

    gf.input_date_by_locale(textbox_start, textbox_end, translate_date_begin, translate_date_end, translate_date_begin_lt, translate_date_end_lt)

    time.sleep(1)
    combobox_order.click()
    time.sleep(1)
    check_if_by_group(ui_self, scale)
    combobox_type.click()
    time.sleep(1)
    gf.locate_and_click(scale + "detali-sum-prekes.png", 0.8)
    time.sleep(1)
    gf.refresh_data(scale)

    sales_export("VGA", "kk/vga", "", "", main_process, export_path, scale)
    gf.print_ui("Vaizdo plokščių (VGA) pardavimai eksportuoti...", ui_self)
    sales_export("MB", "kk/mb", "", "", main_process, export_path, scale)
    gf.print_ui("Pagrindinių plokščių (MB) pardavimai eksportuoti...", ui_self)
    sales_export("SPEC", "", "", "spec", main_process, export_path, scale)
    gf.print_ui("Kiti SPEC pardavimai eksportuoti...", ui_self)
    sales_dlg.close()
    gf.print_ui("Pabaigta eksportuoti pardavimus.", ui_self)

    # Step 3: Exports sale in
    gf.print_ui("Pradedama eksportuoti pirkimus iš įmonės programos...", ui_self)
    gf.navigate_menu(ui_self, scale + "ataskaitos.png", 0.8, scale + "zurnalai.png", 0.8, scale + "prekyba.png", 0.8, main_dlg)

    gf.locate_and_click(scale + "pirkimu-knyga.png", 0.8)

    
    # check_dropdown = pyautogui.locateOnScreen(scale + "pirkimu-knyga.png", confidence=0.8)
    # if check_dropdown == None:
    #     time.sleep(4)
    #     gf.locate_and_click(scale + "pirkimu-knyga.png", 0.8)
    # else:
    #     gf.locate_and_click(scale + "pirkimu-knyga.png", 0.8)
    
    sale_in_dlg = main_dlg.child_window(class_name="TForm_PirkimuKnyga")
    sale_in_dlg.set_focus()
    sale_in_dlg.minimize()
    sale_in_dlg.maximize()
    textbox_start = sale_in_dlg.Edit9
    textbox_end = sale_in_dlg.Edit8

    gf.input_date_by_locale(textbox_start, textbox_end, translate_date_begin, translate_date_end, translate_date_begin_lt, translate_date_end_lt)
    
    op_type = sale_in_dlg.TRxDBLookupCombo2
    op_type.click()
    send_keys("{DOWN}")
    send_keys("{DOWN}")
    send_keys("{DOWN}")
    send_keys("{ENTER}")
    gf.refresh_data(scale)

    vga_wb = Workbook()
    for vga in vga_filter_list:
        export_without_group("vga_" + vga, "", vga, main_process, export_path, scale)
        clipboard_to_excel_without_save(vga_wb)
    vga_wb.save(export_path + "\\" + "VGA_sale_in.xlsx")
    gf.print_ui("Vaizdo plokščių (VGA) pirkimai eksportuoti...", ui_self)
    vga_wb.close()
    sale_in_append(export_path, "VGA_sale_in.xlsx")
    
    mb_wb = Workbook()
    for mb in mb_filter_list:
        export_without_group("mb_" + mb, "", mb, main_process, export_path, scale)
        clipboard_to_excel_without_save(mb_wb)
    mb_wb.save(export_path + "\\" + "mb_sale_in.xlsx")
    gf.print_ui("Pagrindinės plokštės (MB) pirkimai eksportuoti...", ui_self)
    mb_wb.close()
    sale_in_append(export_path, "MB_sale_in.xlsx")


    spec_wb = Workbook()
    export_without_group("spec_", "", "spec", main_process, export_path, scale)
    clipboard_to_excel_without_save(spec_wb)
    spec_wb.save(export_path + "\\" + "spec_sale_in.xlsx")
    gf.print_ui("Kiti SPEC pirkimai eksportuoti...", ui_self)
    spec_wb.close()
    sale_in_append(export_path, "SPEC_sale_in.xlsx")

    

    sale_in_dlg.close()
    gf.print_ui("Pabaigta eksportuoti pirkimus.", ui_self)
    
    # Step 4: Exports stock
    gf.print_ui("Pradedama eksportuoti SPEC likučius iš įmonės programos...", ui_self)
    gf.navigate_menu(ui_self, scale + "ataskaitos.png", 0.8, scale + "vertybiu-likuciai.png", 0.8, scale + "faktiniai-likuciai.png", 0.8, main_dlg)
    stock_dlg = main_dlg.child_window(class_name="TFormEinamiejiVertybiuLikuciai")
    stock_dlg.set_focus()
    stock_dlg.minimize()
    stock_dlg.maximize()
    gf.refresh_data(scale)

    stock_export("VGA", "kk/vga", "", "", main_process, export_path, scale)
    gf.print_ui("Vaizdo plokščių (VGA) likučiai eksportuoti...", ui_self)

    stock_export("MB", "kk/mb", "", "", main_process, export_path, scale)
    gf.print_ui("Pagrindinių plokščių (MB) likučiai eksportuoti...", ui_self)

    stock_export("SPEC", "", "", "spec", main_process, export_path, scale)
    gf.print_ui("Kiti SPEC likučiai eksportuoti...", ui_self)

    stock_dlg.close()
    gf.print_ui("Pabaigta eksportuoti likučius.", ui_self)

    # Step 5: Exports PC sales (Žaliavos) from the program
    gf.print_ui("Pradedama eksportuoti žaliavas iš įmonės programos...", ui_self)
    gf.navigate_menu(ui_self, scale + "ataskaitos.png", 0.8, scale + "gamybos-analize.png", 0.9, scale + "zaliavu-panaudojimas-gaminiuose.png", 0.8, main_dlg)
    pc_sales_dlg = main_dlg.child_window(class_name="TFormZaliavuPanaudojimasGaminiuose")
    pc_sales_dlg.set_focus()
    time.sleep(2)
    pc_sales_dlg.minimize()
    pc_sales_dlg.maximize()
    textbox_start = pc_sales_dlg.Edit14
    textbox_end = pc_sales_dlg.Edit13
    combobox_type = pc_sales_dlg.Combobox
    combobox_group = pc_sales_dlg.TRxDBLookupCombo2

    gf.input_date_by_locale(textbox_start, textbox_end, translate_date_begin, translate_date_end, translate_date_begin_lt, translate_date_end_lt)

    time.sleep(1)
    combobox_group.click()
    time.sleep(1)
    gf.locate_and_click(scale + "grupavimo-nera.png", 0.8)
    time.sleep(1)
    gf.refresh_data(scale)

    vga_wb = Workbook()
    for vga in vga_filter_list:
        pc_sales_export("vga_" + vga, "", vga, main_process, export_path, scale)  
        clipboard_to_excel_without_save(vga_wb)
    
    vga_wb.save(export_path + "\\" + "VGA_with_PC_sale_out.xlsx")
    gf.print_ui("Vaizdo plokščių (VGA) žaliavos eksportuotos...", ui_self)
    vga_wb.close()

    mb_wb = Workbook()
    for mb in mb_filter_list:
        pc_sales_export("mb_" + mb, "", mb, main_process, export_path, scale)
        clipboard_to_excel_without_save(mb_wb)
    mb_wb.save(export_path + "\\" + "MB_with_PC_sale_out.xlsx")
    mb_wb.close()
    gf.print_ui("Pagrindinių plokščių (MB) žaliavos eksportuotos...", ui_self)
    
    pc_sales_export("spec_", "", "spec", main_process, export_path, scale)
    gf.clipboard_to_excel("SPEC", "with_PC_sale_out", export_path)
    gf.print_ui("Kitos SPEC žaliavos eksportuotos...", ui_self)
    
    pc_sales_dlg.close()
    gf.print_ui("Pabaigta eksportuoti žaliavas.", ui_self)
    
    gf.print_ui("Visi duomenys iš įmonės programos eksportuoti sėkmingai.", ui_self)

    # Step 6: Create Report

    gf.print_ui("Visa eksportuota informacija perdaroma į bendra ataskaita...", ui_self)

    prior_month_end = datetime.now().replace(day=1) - timedelta(days=1)
    prior_month_start = prior_month_end.replace(day=1)
    prior_month_start = prior_month_start.strftime("%m.%d")
    prior_month_end = prior_month_end.strftime("%m.%d")
    sale_date = prior_month_start + "-" + prior_month_end
    stock_date = datetime.now().strftime("%m.%d")

    wb_obj = Workbook()

    export_to_sheet(wb_obj, "VGA_sale_in", sale_date, stock_date, export_path, manufacturers, vga_filter_list, mb_filter_list, psu_filter_list, cooler_filter_list, case_filter_list, ssd_filter_list)
    export_to_sheet(wb_obj, "VGA_sale_out", sale_date, stock_date, export_path, manufacturers, vga_filter_list, mb_filter_list, psu_filter_list, cooler_filter_list, case_filter_list, ssd_filter_list)
    export_to_sheet(wb_obj, "VGA_with_PC_sale_out", sale_date, stock_date, export_path, manufacturers, vga_filter_list, mb_filter_list, psu_filter_list, cooler_filter_list, case_filter_list, ssd_filter_list)
    export_to_sheet(wb_obj, "VGA_stock", sale_date, stock_date, export_path, manufacturers, vga_filter_list, mb_filter_list, psu_filter_list, cooler_filter_list, case_filter_list, ssd_filter_list)
    export_to_sheet(wb_obj, "MB_sale_in", sale_date, stock_date, export_path, manufacturers, vga_filter_list, mb_filter_list, psu_filter_list, cooler_filter_list, case_filter_list, ssd_filter_list)
    export_to_sheet(wb_obj, "MB_sale_out", sale_date, stock_date, export_path, manufacturers, vga_filter_list, mb_filter_list, psu_filter_list, cooler_filter_list, case_filter_list, ssd_filter_list)
    export_to_sheet(wb_obj, "MB_with_PC_sale_out", sale_date, stock_date, export_path, manufacturers, vga_filter_list, mb_filter_list, psu_filter_list, cooler_filter_list, case_filter_list, ssd_filter_list)
    export_to_sheet(wb_obj, "MB_stock", sale_date, stock_date, export_path, manufacturers, vga_filter_list, mb_filter_list, psu_filter_list, cooler_filter_list, case_filter_list, ssd_filter_list)
    export_to_sheet(wb_obj, "SPEC_sale_in", sale_date, stock_date, export_path, manufacturers, vga_filter_list, mb_filter_list, psu_filter_list, cooler_filter_list, case_filter_list, ssd_filter_list)
    export_to_sheet(wb_obj, "SPEC_sale_out", sale_date, stock_date, export_path, manufacturers, vga_filter_list, mb_filter_list, psu_filter_list, cooler_filter_list, case_filter_list, ssd_filter_list)
    export_to_sheet(wb_obj, "SPEC_with_PC_sale_out", sale_date, stock_date, export_path, manufacturers, vga_filter_list, mb_filter_list, psu_filter_list, cooler_filter_list, case_filter_list, ssd_filter_list)
    export_to_sheet(wb_obj, "SPEC_stock", sale_date, stock_date, export_path, manufacturers, vga_filter_list, mb_filter_list, psu_filter_list, cooler_filter_list, case_filter_list, ssd_filter_list)

    if 'Sheet' in wb_obj.sheetnames:
        wb_obj.remove(wb_obj['Sheet'])
    wb_obj.save(export_parent_directory + "\\" + "SPEC_" + today_date + ".xlsx")
    wb_obj.close()


    


    # finishing work

    end_time = time.time()
    seconds_spent_executing = end_time - start_time
    executing_time = timedelta(seconds=seconds_spent_executing)
    gf.print_ui("Visi SPEC mėnesinės ataskaitos modulio darbai pabaigti per: [" + str(executing_time) + "]", ui_self)

if __name__ == "__main__":
    main(None, "", "")